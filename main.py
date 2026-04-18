import sys

import numpy as np
from PyQt6 import uic
from datetime import datetime
from PyQt6.QtCore import QSize, pyqtSignal, QTimer
from PyQt6.QtGui import QIcon
from PyQt6.QtWidgets import QMainWindow, QTableWidgetItem, QStyle, QMessageBox, QTableWidget, QFileDialog
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout,
    QPushButton, QLabel
)
from PyQt6.QtGui import QColor, QBrush

from Parser.parser import parse_1c_table, pre_parser
from database import Database
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import qdarkstyle



style = QApplication.style()
LAST_WINDOW_POS = None

def center_window(window):
    global LAST_WINDOW_POS

    geo = window.frameGeometry()

    if LAST_WINDOW_POS:
        # 👉 двигаем центр окна в сохранённую точку
        geo.moveCenter(LAST_WINDOW_POS)
        window.move(geo.topLeft())
        return

    # 👉 стандартный центр экрана
    screen = window.screen()
    screen_geometry = screen.availableGeometry()

    geo.moveCenter(screen_geometry.center())
    window.move(geo.topLeft())

class BaseWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self._first_show = True

    def showEvent(self, event):
        super().showEvent(event)

        if self._first_show:
            center_window(self)
            self._first_show = False

    def moveEvent(self, event):
        super().moveEvent(event)

        global LAST_WINDOW_POS

        if not self._first_show:
            geo = self.frameGeometry()
            LAST_WINDOW_POS = geo.center()  # 👉 сохраняем центр

class BaseWidget(QWidget):
    def __init__(self):
        super().__init__()
        self._first_show = True

    def showEvent(self, event):
        super().showEvent(event)

        if self._first_show:
            center_window(self)
            self._first_show = False

    def moveEvent(self, event):
        super().moveEvent(event)

        global LAST_WINDOW_POS

        if not self._first_show:
            geo = self.frameGeometry()
            LAST_WINDOW_POS = geo.center()  # 👉 сохраняем центр

class MainWindow(BaseWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi("gui/main.ui", self)
        center_window(self)

        # подключаем кнопку
        self.btn_start.clicked.connect(self.open_firm_window)
        self.btn_exit.clicked.connect(self.close)

        self.path  = self.text_db.text().strip()

        self.setFixedSize(self.size())

    def open_firm_window(self):
        self.firm_window = FirmWindow(self.path)
        self.firm_window.show()

        self.close()  # закрываем текущее окно


class Edit_Firm(BaseWidget):

    firm_edit = pyqtSignal()


    def __init__(self, db, id):
        super().__init__()
        uic.loadUi("gui/edit_firm.ui", self)
        center_window(self)

        self.id = id

        name = db.get_firm(self.id)
        self.name = name["name"]

        self.text_edit_firm.setText(self.name)

        self.db = db
        # подключаем кнопку
        self.btn_edit_firm.clicked.connect(self.edit_firm)

        self.btn_exit_edit_firm.clicked.connect(self.close)

        self.setFixedSize(self.size())


    def edit_firm(self):

        new_firm = self.text_edit_firm.text().strip()

        if not new_firm:
            QMessageBox.warning(
                self,
                "Ошибка",
                "Поле с названием пустое. Добавление отменено"
            )
            return

        success, error = self.db.update_firm(self.id, new_firm)

        if success:
            self.firm_edit.emit()
            self.close()
        else:
            QMessageBox.warning(
                self,
                "Ошибка",
                "Такое название уже существует"
            )


class Add_Cheks(BaseWidget):

    checks_add = pyqtSignal()


    def __init__(self, db, firm_id):
        super().__init__()
        uic.loadUi("gui/create_checks.ui", self)
        center_window(self)

        self.db = db

        self.firm_id = firm_id

        self.spin_materiality.setMinimum(0)
        self.spin_materiality.setMaximum(1_000_000_000)
        self.spin_materiality.setDecimals(2)

        # подключаем кнопку
        self.btn_checks_create.clicked.connect(self.add)

        self.btn_checks_back.clicked.connect(self.close)

        self.setFixedSize(self.size())


    def add(self):

        period  = self.text_period_checks.text().strip()

        materiality = self.spin_materiality.value()

        if not period or not materiality:
            QMessageBox.warning(
                self,
                "Ошибка",
                "Не все поля заполнены"
            )
            return

        self.db.add_check(self.firm_id, period, materiality)
        self.checks_add.emit()
        self.close()

class Edit_Checks(BaseWidget):

    check_edit = pyqtSignal()


    def __init__(self, db, check_id):
        super().__init__()
        uic.loadUi("gui/edit_checks.ui", self)
        center_window(self)

        self.db = db

        self.check_id = check_id

        self.spin_materiality_edit.setMinimum(0)
        self.spin_materiality_edit.setMaximum(1_000_000_000)
        self.spin_materiality_edit.setDecimals(2)

        temp = self.db.get_one_checks(self.check_id)

        self.period = temp["period"][0]
        self.materiality = temp["materiality"][0]
        self.created_at = temp["created_at"][0]

        self.text_period_checks_edit.setText(self.period)

        self.spin_materiality_edit.setValue(self.materiality)

        self.btn_checks_edit_back.clicked.connect(self.close)
        # подключаем кнопку
        self.btn_checks_create_edit.clicked.connect(self.edit)

        self.setFixedSize(self.size())


    def edit(self):

        period  = self.text_period_checks_edit.text().strip()

        materiality = self.spin_materiality_edit.value()

        if not period or not materiality:
            QMessageBox.warning(
                self,
                "Ошибка",
                "Не все поля заполнены!"
            )
            return

        self.db.update_check(self.check_id, period, materiality)
        self.check_edit.emit()
        self.close()


class ResultWindow(BaseWindow):

    def __init__(self, db, firm_id, check_id, acc_id):
        super().__init__()
        uic.loadUi("gui/results.ui", self)
        center_window(self)

        self.db = db

        self.firm_id = firm_id
        self.check_id = check_id

        self.acc_id = acc_id

        temp = self.db.get_accounts_byid(self.acc_id)
        name = temp['name'][0]


        # подключаем кнопку
        self.text_result_name.setText(name)

        self.load_preview_table()

        self.btn_result_back.clicked.connect(self.open_data)

        self.btn_calc_result.clicked.connect(self.calc_result)

        self.btn_save_xmls.clicked.connect(self.save_excel)

        self.setFixedSize(self.size())

    def calc_result(self):

        self.db.clear_results(self.acc_id)
        self.db.run_mus(self.check_id, self.acc_id)
        self.load_preview_table()

    def open_data(self):
        self.close()
        # тут откроешь окно проверок
        self.open_data_win = DataWindow(self.db, self.firm_id, self.check_id, self.acc_id)
        self.open_data_win.show()

    def save_excel(self):
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить файл",
            "",
            "Excel Files (*.xlsx)"
        )

        if path:
            if not path.endswith(".xlsx"):
                path += ".xlsx"

            self.export_to_excel(path)

    def export_to_excel(self, path):


        # --- 1. параметры
        params = [
            ("PM", self.PM),
            ("n", self.n),
            ("h", self.h),
            ("Покрытие (%)", round(self.coverage, 2)),
            ("Итого", self.total),
            ("High Value", self.high_value_sum),
            ("MUS", self.mus_sum),
            ("Test Sum", self.test_sum),
            ("Сообщения", self.messages)
        ]

        # --- 2. выборка
        df = self.df.copy()
        df.columns = [
            "Номер документа",
            "Дата",
            "Примечание",
            "Сумма",
            "Причина"
        ]

        # --- 3. запись через pandas
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="MUS", index=False, startrow=12)

        # --- 4. открываем для форматирования
        wb = load_workbook(path)
        ws = wb["MUS"]

        # --- 🎯 Стили
        bold = Font(bold=True, size=12)
        header_font = Font(bold=True)
        center = Alignment(vertical="center")
        wrap = Alignment(wrap_text=True, vertical="top")

        fill_high = PatternFill("solid", fgColor="C6EFCE")  # зелёный
        fill_mus = PatternFill("solid", fgColor="FFF2CC")  # жёлтый

        # --- 5. Заголовок
        ws["A1"] = "Результаты MUS выборки"
        ws["A1"].font = Font(bold=True, size=14)

        # --- 6. Параметры
        row = 3
        for name, value in params:
            ws[f"A{row}"] = name
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = bold
            row += 1

        # --- 7. Шапка таблицы
        header_row = 13

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = header_font
            cell.alignment = center

        # --- 8. Подсветка строк
        for row in range(header_row + 1, ws.max_row + 1):
            reason = ws.cell(row=row, column=5).value  # "Причина"

            if reason == "HIGH_VALUE":
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill_high

            elif reason == "MUS":
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill_mus

        # --- 9. Перенос текста (примечание)
        for row in range(header_row + 1, ws.max_row + 1):
            ws.cell(row=row, column=3).alignment = wrap  # Примечание

        # --- 10. Автоширина колонок
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # --- 11. Закрепляем шапку
        ws.freeze_panes = "A14"

        # --- 12. Фильтр
        ws.auto_filter.ref = f"A{header_row}:E{ws.max_row}"

        wb.save(path)


    def load_preview_table(self):

        temp = self.db.get_results(self.acc_id)

        if temp:


            self.PM = temp['PM']
            self.n = temp['n']
            self.h = temp['h']
            self.coverage = temp['coverage']
            self.total = temp['total']
            self. high_value_sum= temp['high_value_sum']
            self.mus_sum = temp['mus_sum']
            self.test_sum= temp['test_sum']
            self.messages = temp['messages']
            self.df= temp['sample']


            self.table_results.setRowCount(self.df.shape[0])
            self.table_results.setColumnCount(self.df.shape[1])

            # заголовки
            self.table_results.setHorizontalHeaderLabels([
                "Номер документа",
                "Дата",
                "Примечание",
                "Сумма",
                "Причина"
            ])

            self.table_results.setColumnWidth(0, 150)
            self.table_results.setColumnWidth(1, 150)
            self.table_results.setColumnWidth(2, 300)
            self.table_results.setColumnWidth(3, 150)
            self.table_results.setColumnWidth(4, 150)

            # заполнение
            for row_idx in range(self.df.shape[0]):
                for col_idx in range(self.df.shape[1]):
                    value = self.df.iat[row_idx, col_idx]

                    item = QTableWidgetItem(str(value))
                    self.table_results.setItem(row_idx, col_idx, item)

            # read-only
            self.table_results.setEditTriggers(
                QTableWidget.EditTrigger.NoEditTriggers
            )
            self.table_results.verticalHeader().setVisible(False)

            layout1 = self.widget_results.layout()
            layout2 = self.widget_results_2.layout()
            layout3 = self.widget_results_3.layout()

            while layout1.rowCount():
                layout1.removeRow(0)
            while layout2.rowCount():
                layout2.removeRow(0)
            while layout3.rowCount():
                layout3.removeRow(0)

            layout1.addRow("PM:", QLabel(str(self.PM)))
            layout1.addRow("n:", QLabel(str(self.n)))
            layout1.addRow("h:", QLabel(str(self.h)))
            layout1.addRow("Покрытие:", QLabel(f"{self.coverage:.2%}"))
            layout2.addRow("Итого:", QLabel(str(self.total)))
            layout2.addRow("Наибольшая сумма:", QLabel(str(self.high_value_sum)))
            layout2.addRow("MUS сумма:", QLabel(str(self.mus_sum)))
            layout2.addRow("Тестовая сумма:", QLabel(str(self.test_sum)))

            msg_label = QLabel("".join(self.messages))
            msg_label.setWordWrap(True)
            msg_label.setStyleSheet("color: red;")

            layout3.addRow("Сообщение:", msg_label)


class DataWindow(BaseWindow):


    def __init__(self, db, firm_id,check_id, acc_id):
        super().__init__()
        uic.loadUi("gui/data.ui", self)
        center_window(self)

        self.db = db

        self.firm_id = firm_id
        self.check_id = check_id

        self.acc_id = acc_id

        temp = self.db.get_accounts_byid(self.acc_id)
        name = temp['name'][0]
        # подключаем кнопку
        self.text_data_account_name.setText(name)

        self.btn_data_exit_acc.clicked.connect(self.open_acc)

        self.btn_raw_entity.clicked.connect(self.open_entity)

        self.btn_result.clicked.connect(self.open_result)

        self.setFixedSize(self.size())

    def open_result(self):
        self.close()
        # тут откроешь окно проверок
        self.open_result_win = ResultWindow(self.db, self.firm_id, self.check_id, self.acc_id)
        self.open_result_win.show()


    def open_acc(self):
        self.close()
        # тут откроешь окно проверок
        self.open_checks = AccountWindow(self.db, self.firm_id, self.check_id)
        self.open_checks.show()

    def open_entity(self):
        self.close()
        # тут откроешь окно проверок
        self.open_entity_win = EntityWindow(self.db, self.firm_id, self.check_id, self.acc_id)
        self.open_entity_win.show()


class RawData(BaseWindow):

    def __init__(self, db, firm_id, check_id, acc_id, db_raw):
        super().__init__()
        uic.loadUi("gui/raw.ui", self)
        center_window(self)

        self.db = db
        self.firm_id = firm_id
        self.check_id = check_id
        self.acc_id = acc_id

        self.db_raw = db_raw

        temp = self.db.get_accounts_byid(self.acc_id)
        name = temp['name'][0]

        # подключаем кнопку
        self.text_raw_name_acc.setText(name)

        self.load_preview_table(self.db_raw)

        self.btn_raw_back.clicked.connect(self.open_entity)

        self.btn_raw_new_entity.clicked.connect(self.add_new_entity)

        self.btn_raw_preview.clicked.connect(self.apply_preview)

        self.setFixedSize(self.size())

    def add_new_entity(self):

        try:
            first_row = int(self.text_first_row.text())
            last_row = int(self.text_last_row.text())



            doc_col = int(self.text_doc_id_col.text())
            date_col = int(self.text_data_col.text())

            amount_col = int(self.text_raw_amount.text())


            note_cols = [
                int(x.strip())
                for x in self.text_note_list.text().split(",")
                if x.strip()
            ]

            result_df= parse_1c_table(
                self.db_raw,
                id_col=doc_col,
                date_col=date_col,
                amount_col=amount_col,
                start_row=first_row,
                end_row=last_row,
                extra_cols=note_cols
            )

            self.db.insert_entries(self.acc_id, result_df)

            self.open_entity()

        except:
            QMessageBox.warning(self, "Ошибка", "Проверьте введённые значения")

    def clear_colors(self, table):
        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item:
                    item.setBackground(QBrush())  # ✅ сброс
                    item.setForeground(QBrush())  # (если вдруг тоже менял)

    def apply_preview(self):

        self.clear_colors(self.table_raw_start)
        self.clear_colors(self.table_raw_end)

        try:
            first_row = int(self.text_first_row.text())
            last_row = int(self.text_last_row.text())



            doc_col = int(self.text_doc_id_col.text())
            date_col = int(self.text_data_col.text())

            amount_col = int(self.text_raw_amount.text())


            note_cols = [
                int(x.strip())
                for x in self.text_note_list.text().split(",")
                if x.strip()
            ]

            self.highlight_table(
                self.table_raw_start,
                self.head,
                first_row,
                last_row,
                {
                    "doc": doc_col,
                    "date": date_col,
                    "amount": amount_col,
                    "note": note_cols
                }
            )

            self.highlight_table(
                self.table_raw_end,
                self.tail,
                first_row,
                last_row,
                {
                    "doc": doc_col,
                    "date": date_col,
                    "amount": amount_col,
                    "note": note_cols
                }
            )
        except:
            QMessageBox.warning(self, "Ошибка", "Проверьте введённые значения")

    def open_entity(self):
        self.close()

        self.open_entity_win = EntityWindow(self.db, self.firm_id, self.check_id, self.acc_id)
        self.open_entity_win.show()

    def load_preview_table(self, df):

        df = df.replace({np.nan: ""})
        self.head= df.head(30)

        self.table_raw_start.setRowCount(self.head.shape[0])
        self.table_raw_start.setColumnCount(self.head.shape[1])

        # заголовки
        self.table_raw_start.setHorizontalHeaderLabels(
            [str(col) for col in self.head.columns]
        )

        # заполнение
        for row_idx in range(self.head.shape[0]):
            for col_idx in range(self.head.shape[1]):
                value = self.head.iat[row_idx, col_idx]

                item = QTableWidgetItem(str(value))
                self.table_raw_start.setItem(row_idx, col_idx, item)

        # read-only
        self.table_raw_start.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers
        )
        self.table_raw_start.verticalHeader().setVisible(False)

        self.tail = df.tail(30)

        self.table_raw_end.setRowCount(self.tail.shape[0])
        self.table_raw_end.setColumnCount(self.tail.shape[1])

        # заголовки
        self.table_raw_end.setHorizontalHeaderLabels(
            [str(col) for col in self.tail.columns]
        )

        # заполнение
        for row_idx in range(self.tail.shape[0]):
            for col_idx in range(self.tail.shape[1]):
                value = self.tail.iat[row_idx, col_idx]

                item = QTableWidgetItem(str(value))
                self.table_raw_end.setItem(row_idx, col_idx, item)

        # read-only
        self.table_raw_end.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers
        )
        self.table_raw_start.verticalHeader().setVisible(False)
        self.table_raw_end.verticalHeader().setVisible(False)

        if len(df) <= 30:
            self.table_raw_end.hide()
        else:
            self.table_raw_end.show()

    def highlight_table(self, table, df_part, first_row, last_row, cols_map):

        col_positions = {
            col_name: idx
            for idx, col_name in enumerate(df_part.columns)
        }

        note_cols = cols_map.get("note", [])

        # 👉 цвет текста для тёмной темы
        text_color = QColor(220, 220, 220)

        for row_idx in range(df_part.shape[0]):
            real_row = df_part.iloc[row_idx, 0]

            if not (first_row <= real_row <= last_row):
                continue

            # --- doc/date/amount
            for key, color in [
                ("doc", QColor(80, 120, 80)),  # более тёмные цвета
                ("date", QColor(120, 120, 60)),
                ("amount", QColor(140, 100, 60)),
            ]:
                col_name = cols_map.get(key)

                if col_name in col_positions:
                    col_idx = col_positions[col_name]

                    item = table.item(row_idx, col_idx)
                    if item:
                        item.setBackground(QBrush(color))
                        item.setForeground(QBrush(text_color))  # 👈 фикс

            # --- note
            for col_name in note_cols:
                if col_name in col_positions:
                    col_idx = col_positions[col_name]

                    item = table.item(row_idx, col_idx)
                    if item:
                        item.setBackground(QBrush(QColor(100, 80, 140)))
                        item.setForeground(QBrush(text_color))  # 👈 фикс

class EntityWindow(BaseWindow):

    def __init__(self, db, firm_id, check_id, acc_id):
        super().__init__()
        uic.loadUi("gui/entity.ui", self)
        center_window(self)

        self.db = db

        self.firm_id = firm_id
        self.check_id = check_id

        self.acc_id = acc_id

        temp = self.db.get_accounts_byid(self.acc_id)
        name = temp['name'][0]


        # подключаем кнопку
        self.text_entity_account_name.setText(name)

        self.load_entity()

        self.btn_entity_exit_to_data.clicked.connect(self.open_data)

        self.btn_entity_new_file.clicked.connect(self.load_file)

        self.setFixedSize(self.size())


    def load_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите файл",
            "",
            "Excel Files (*.xlsx *.xls *.xlsm *.csv)"
        )

        if file_path:
            df_raw = pre_parser(file_path)
            self.close()
            self.open_new_entity = RawData(self.db, self.firm_id, self.check_id, self.acc_id, df_raw)

            self.open_new_entity.show()


    def open_data(self):
        self.close()
        # тут откроешь окно проверок
        self.open_data_win = DataWindow(self.db, self.firm_id, self.check_id, self.acc_id)
        self.open_data_win.show()

    def load_entity(self):

        entries = self.db.get_entries(self.acc_id)

        self.table_entitys.setRowCount(len(entries))
        self.table_entitys.setColumnCount(5)

        self.table_entitys.setHorizontalHeaderLabels([
            "", "Документ", "Дата", "Сумма", "Примечание"
        ])

        self.table_entitys.setColumnWidth(1, 155)
        self.table_entitys.setColumnWidth(2, 110)
        self.table_entitys.setColumnWidth(3, 110)
        self.table_entitys.setColumnWidth(4, 300)


        for row_idx, (id, account_id, doc_id, date, amount, note) in enumerate(entries):
            # --- ID (скрытый)
            self.table_entitys.setItem(
                row_idx, 0,
                QTableWidgetItem(str(id))
            )

            # --- Документ
            self.table_entitys.setItem(
                row_idx, 1,
                QTableWidgetItem(doc_id)
            )

            # --- Дата
            self.table_entitys.setItem(
                row_idx, 2,
                QTableWidgetItem(str(date))
            )

            # --- Сумма
            self.table_entitys.setItem(
                row_idx, 3,
                QTableWidgetItem(str(amount))
            )

            # --- Примечание
            self.table_entitys.setItem(
                row_idx, 4,
                QTableWidgetItem(note)
            )


        # 👉 скрываем колонку ID
        self.table_entitys.setColumnHidden(0, True)

        self.table_entitys.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers
        )



class Edit_Acc(BaseWidget):

    edit_acc = pyqtSignal()

    def __init__(self, db, acc_id):
        super().__init__()
        uic.loadUi("gui/edit_acc.ui", self)
        center_window(self)

        self.db = db

        self.acc_id = acc_id

        temp  = self.db.get_accounts_byid(self.acc_id)
        name = temp['name'][0]
        # подключаем кнопку
        self.text_account_edit_name.setText(name)

        self.btn_account_edit.clicked.connect(self.edit_account)

        self.btn_account_edit_exit.clicked.connect(self.close)

        self.setFixedSize(self.size())


    def edit_account(self):
        new_acc_name = self.text_account_edit_name.text().strip()

        if not new_acc_name:
            QMessageBox.warning(
                self,
                "Ошибка",
                "Поле с названием пустое. Добавление отменено"
            )
            return

        self.db.update_acc(id=self.acc_id, name=new_acc_name)

        self.edit_acc.emit()
        self.close()

class Create_acc(BaseWindow):

    acc_added = pyqtSignal()

    def __init__(self, db, firm_id, check_id):
        super().__init__()
        uic.loadUi("gui/create_acc.ui", self)
        center_window(self)

        self.db = db
        self.firm_id = firm_id

        self.check_id = check_id
        # подключаем кнопку
        self.btn_account_create.clicked.connect(self.create_account)

        self.btn_account_create_back_checks.clicked.connect(self.back_account)

        self.open_checks = AccountWindow(self.db, self.firm_id, check_id)

        self.setFixedSize(self.size())

    def back_account(self):

        self.close()


    def create_account(self):
        new_acc_name = self.text_account_name.text().strip()

        if not new_acc_name:
            QMessageBox.warning(
                self,
                "Ошибка",
                "Поле с названием пустое. Добавление отменено"
            )
            return

        new_id, created= self.db.get_or_create_account(check_id=self.check_id, name=new_acc_name)

        if created:
            self.acc_added.emit()
            self.close()
        else:
            QMessageBox.warning(
                self,
                "Ошибка",
                "Такой счет уже существует"
            )


class AccountWindow(BaseWindow):

    def __init__(self, db, firm_id, check_id):
        super().__init__()
        uic.loadUi("gui/account.ui", self)
        center_window(self)

        self.db = db
        self.firm_id = firm_id
        self.check_id = check_id

        name = db.get_firm(self.firm_id)
        self.firm_name = name["name"]

        self.text_account_name_firm.setText(self.firm_name)

        temp = self.db.get_one_checks(self.check_id)

        self.period = temp["period"][0]

        self.text_account_checks_name.setText(self.period)

        self.load_accounts()
        # подключаем кнопку

        self.btn_new_account.clicked.connect(self.open_create_account)

        self.table_accounts.cellDoubleClicked.connect(self.open_data)

        self.btn_accounts_exit_to_checks.clicked.connect(self.open_checks)

        self.setFixedSize(self.size())


    def open_create_account(self):
        self.create_account_win = Create_acc(self.db, self.firm_id, self.check_id)
        self.create_account_win.acc_added.connect(self.load_accounts)
        self.create_account_win.show()

    def open_data(self, row, column):

        acc_id_item = self.table_accounts.item(row, 0)

        if acc_id_item:
            acc_id = int(acc_id_item.text())

            self.close()
            # тут откроешь окно проверок
            self.open_data_win = DataWindow(self.db, self.firm_id, self.check_id, acc_id)
            self.open_data_win.show()


    def open_checks(self):

        self.close()
        self.open = CheckWindow(firm_id= self.firm_id, db =self.db)
        self.open.show()

    def edit_account(self, acc_id):
        self.edit = Edit_Acc(self.db, acc_id)
        self.edit.edit_acc.connect(self.load_accounts)
        self.edit.show()

    def delete_acc(self, acc_id):
        reply = QMessageBox.question(
            self,
            "Удаление",
            "Удалить проверку и ВСЕ связанные данные?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply != QMessageBox.StandardButton.Yes:
            return
        else:
            self.db.delete_accounts(acc_id)
            self.load_accounts()

    def load_accounts(self):


        accounts = self.db.get_accounts(self.check_id)

        self.table_accounts.setRowCount(len(accounts))
        self.table_accounts.setColumnCount(4)

        self.table_accounts.setHorizontalHeaderLabels([
            "", "Название", "", ""
        ])

        self.table_accounts.setColumnWidth(1, 470)
        self.table_accounts.setColumnWidth(2, 30)
        self.table_accounts.setColumnWidth(3, 30)

        for row_idx, (id, name) in enumerate(accounts):
            # --- ID (скрытый)
            self.table_accounts.setItem(
                row_idx, 0,
                QTableWidgetItem(str(id))
            )

            # --- нзвание
            self.table_accounts.setItem(
                row_idx, 1,
                QTableWidgetItem(name)
            )

            btn_account_edit = QPushButton()
            btn_account_edit.setIcon(QIcon("gui/icons/edit.png"))
            btn_account_edit.clicked.connect(lambda _, fid=id: self.edit_account(fid))

            btn_account_del = QPushButton()
            btn_account_del.setIcon(QIcon("gui/icons/trash.png"))
            btn_account_del.clicked.connect(lambda _, fid=id: self.delete_acc(fid))

            btn_account_edit.setFixedSize(30, 30)
            btn_account_del.setFixedSize(30, 30)

            btn_account_edit.setIconSize(QSize(30, 30))
            btn_account_del.setIconSize(QSize(30, 30))


            self.table_accounts.setCellWidget(row_idx, 2, btn_account_edit)
            self.table_accounts.setCellWidget(row_idx, 3, btn_account_del)

        # 👉 скрываем колонку ID
        self.table_accounts.setColumnHidden(0, True)

        self.table_accounts.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers
        )

class CheckWindow(BaseWindow):

    def __init__(self, db, firm_id):
        super().__init__()
        uic.loadUi("gui/checks.ui", self)
        center_window(self)

        self.db = db
        self.firm_id = firm_id


        name = db.get_firm(self.firm_id )
        self.name = name["name"]

        self.text_name_firm.setText(self.name)

        self.load_checks()
        # подключаем кнопку
        self.btn_new_checks.clicked.connect(self.open_create_cheks)

        self.table_checks.cellDoubleClicked.connect(self.open_account)

        self.btn_checks_exit_to_firm.clicked.connect(self.open_firm)

        self.setFixedSize(self.size())


    def open_account(self, row, column):
        check_id_item = self.table_checks.item(row, 0)

        if check_id_item:
            check_id = int(check_id_item.text())

            self.close()
            # тут откроешь окно проверок
            self.open_checks = AccountWindow(self.db, self.firm_id, check_id)
            self.open_checks.show()


    def open_firm(self):

        self.close()
        self.firm_window = FirmWindow(db =self.db)
        self.firm_window.show()

    def open_create_cheks(self):
        self.create_cheks = Add_Cheks(self.db, self.firm_id)
        self.create_cheks.checks_add.connect(self.load_checks)
        self.create_cheks.show()

    def edit_checks(self, check_id):
        self.edit = Edit_Checks(self.db, check_id)
        self.edit.check_edit.connect(self.load_checks)
        self.edit.show()

    def delete_cheks(self, check_id):
        reply = QMessageBox.question(
            self,
            "Удаление",
            "Удалить проверку и ВСЕ связанные данные?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply != QMessageBox.StandardButton.Yes:
            return
        else:
            self.db.delete_checks(check_id)
            self.load_checks()


    def load_checks(self):


        checks = self.db.get_checks(self.firm_id)

        self.table_checks.setRowCount(len(checks))
        self.table_checks.setColumnCount(6)

        self.table_checks.setHorizontalHeaderLabels([
            "", "Период", "Уровень сущест.", "Дата создания", "", ""
        ])

        self.table_checks.setColumnWidth(1, 150)
        self.table_checks.setColumnWidth(2, 150)
        self.table_checks.setColumnWidth(3, 150)
        self.table_checks.setColumnWidth(4, 30)
        self.table_checks.setColumnWidth(5, 30)

        for row_idx, (id, period, materiality, created_at) in enumerate(checks):
            # --- ID (скрытый)
            self.table_checks.setItem(
                row_idx, 0,
                QTableWidgetItem(str(id))
            )

            # --- Периорд
            self.table_checks.setItem(
                row_idx, 1,
                QTableWidgetItem(period)
            )

            # --- Уровень
            self.table_checks.setItem(
                row_idx, 2,
                QTableWidgetItem(str(materiality))
            )
            # --- Дата
            self.table_checks.setItem(
                row_idx, 3,
                QTableWidgetItem(str(created_at)[:10])
            )


            btn_checks_edit = QPushButton()
            btn_checks_edit.setIcon(QIcon("gui/icons/edit.png"))
            btn_checks_edit.clicked.connect(lambda _, fid=id: self.edit_checks(fid))

            btn_checks_del = QPushButton()
            btn_checks_del.setIcon(QIcon("gui/icons/trash.png"))
            btn_checks_del.clicked.connect(lambda _, fid=id: self.delete_cheks(fid))

            btn_checks_edit.setFixedSize(30, 30)
            btn_checks_del.setFixedSize(30, 30)

            btn_checks_edit.setIconSize(QSize(30, 30))
            btn_checks_del.setIconSize(QSize(30, 30))


            self.table_checks.setCellWidget(row_idx, 4, btn_checks_edit)
            self.table_checks.setCellWidget(row_idx, 5, btn_checks_del)

        # 👉 скрываем колонку ID
        self.table_checks.setColumnHidden(0, True)

        self.table_checks.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers
        )



class Create_Firm(BaseWindow):

    firm_added = pyqtSignal()

    def __init__(self, db):
        super().__init__()
        uic.loadUi("gui/create_firm.ui", self)
        center_window(self)

        self.db = db
        # подключаем кнопку
        self.btn_create_firm.clicked.connect(self.create_firm)

        self.btn_exit_create_firm.clicked.connect(self.close)

        self.setFixedSize(self.size())

    def create_firm(self):
        new_firm = self.text_new_firm.text().strip()

        if not new_firm:
            QMessageBox.warning(
                self,
                "Ошибка",
                "Поле с названием пустое. Добавление отменено"
            )
            return

        firm_id, created = self.db.get_or_create_firm(new_firm)

        if created:
            self.firm_added.emit()
            self.close()
        else:
            QMessageBox.warning(
                self,
                "Ошибка",
                "Такая фирма уже существует"
            )



class FirmWindow(BaseWindow):
    def __init__(self, path="audit.db", db=None):
        super().__init__()
        uic.loadUi("gui/firm.ui", self)
        center_window(self)

        if db:
            self.db = db
        else:
            self.db = Database(path=path)
            self.db.init_db()


        self.load_firms()
        self.table_firms.cellDoubleClicked.connect(self.open_firm)
        self.btn_new_firm.clicked.connect(self.open_create_firm_window)

        # self.btn_search_firm.clicked.connect(self.search_firm)
        self.search_timer = QTimer()
        self.search_timer.setSingleShot(True)
        self.search_timer.timeout.connect(self.perform_search)

        # реакция на ввод
        self.text_search_firm.textChanged.connect(self.on_search_text_changed)
        self.btn_search_reset.clicked.connect(self.reset_search)

        self.setFixedSize(self.size())


    def reset_search(self):
        self.text_search_firm.clear()

    def on_search_text_changed(self):
        # перезапускаем таймер при каждом вводе
        self.search_timer.start(300)  # 300 мс

    def perform_search(self):
        query = self.text_search_firm.text().strip()
        self.load_firms(query=query)


    def load_firms(self, query=""):

        if query == "":
            firms = self.db.get_firms()
        else:
            firms = self.db.search_firms(query)

        self.table_firms.setRowCount(len(firms))
        self.table_firms.setColumnCount(4)

        self.table_firms.setHorizontalHeaderLabels([
            "", "Название", "", ""
        ])

        self.table_firms.setColumnWidth(1, 470)
        self.table_firms.setColumnWidth(2, 30)
        self.table_firms.setColumnWidth(3, 30)

        for row_idx, (firm_id, name) in enumerate(firms):
            # --- ID (скрытый)
            self.table_firms.setItem(
                row_idx, 0,
                QTableWidgetItem(str(firm_id))
            )

            # --- Название
            self.table_firms.setItem(
                row_idx, 1,
                QTableWidgetItem(name)
            )


            btn_edit = QPushButton()
            btn_edit.setIcon(QIcon("gui/icons/edit.png"))
            btn_edit.clicked.connect(lambda _, fid=firm_id: self.edit_firm(fid))

            btn_del = QPushButton()
            btn_del.setIcon(QIcon("gui/icons/trash.png"))
            btn_del.clicked.connect(lambda _, fid=firm_id: self.delete_firm(fid))

            btn_edit.setFixedSize(30, 30)
            btn_del.setFixedSize(30, 30)

            btn_edit.setIconSize(QSize(30, 30))
            btn_del.setIconSize(QSize(30, 30))


            self.table_firms.setCellWidget(row_idx, 2, btn_edit)
            self.table_firms.setCellWidget(row_idx, 3, btn_del)

        # 👉 скрываем колонку ID
        self.table_firms.setColumnHidden(0, True)

        self.table_firms.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers
        )

    def open_firm(self, row, column):
        firm_id_item = self.table_firms.item(row, 0)

        if firm_id_item:
            firm_id = int(firm_id_item.text())

            self.close()
            # тут откроешь окно проверок
            self.open_checks = CheckWindow(self.db, firm_id)
            self.open_checks.show()


    def edit_firm(self, firm_id):
        self.create_firm = Edit_Firm(self.db, firm_id)
        self.create_firm.firm_edit.connect(self.load_firms)
        self.create_firm.show()

    def delete_firm(self, firm_id):
        reply = QMessageBox.question(
            self,
            "Удаление",
            "Удалить фирму и ВСЕ связанные данные?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply != QMessageBox.StandardButton.Yes:
            return
        else:
            self.db.delete_firm(firm_id)
            self.load_firms()



    def open_create_firm_window(self):
        self.create_firm = Create_Firm(self.db)
        self.create_firm.firm_added.connect(self.load_firms)
        self.create_firm.show()





if __name__ == "__main__":
    app = QApplication(sys.argv)

    app.setStyleSheet(qdarkstyle.load_stylesheet())

    window = MainWindow()
    window.show()

    sys.exit(app.exec())